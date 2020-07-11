"""Examples of how the classes in opt_models can be used.
This is not an actual program.
"""


import opt_models as models
import pandas as pd
from openpyxl import Workbook, load_workbook

# read input data from excel files to DataFrames
product_data = pd.read_excel('product_data.xlsx',index_col=0)
mold_data = pd.read_excel('mold_data.xlsx',index_col=0)
prev_mount = pd.read_excel('mounted.xlsx',index_col=0)['part']
# solver termination conditons (terminate if MIPGap = 0 or time elapsed 100secs)
term_conds = [(0,100)]
# production line parameters
hours, arms, mounts = 120, 2, 2

# build a model to allocate production that maximize profits while meeting demand
model1 = models.MaxProfit(hours,arms,mounts)
model1.read_data(product_data, mold_data)
model1.build_model()
# uncomment the line below to write the model to a .lp file to inspect the formulation
# model1.write('filename.lp')
model1.optimize(term_conds)
model1.update_production()

# get result from profit optimization
produced = model1.get_production()
# concatenate produced to product_data for use in scheduling optimization
product_data = pd.concat((product_data,produced),axis=1)

# build a second model to fit the desired production into a schedule
# that minimizes mold changes
model2 = models.ProductionSchedule(hours,arms,mounts)
model2.read_data(product_data, mold_data,prev_mount)
model2.build_phase1()
# model2.write('filename2.lp')
model2.optimize(term_conds)
model2.update_production()
model2.build_phase2()
# model2.write('filename3.lp')
model2.optimize(term_conds)
model2.update_schedule()

schedule = model2.get_schedule()
# try loading excel file
try:
    wb = load_workbook('production_schedule_1.xlsx')
# if it doesn't exist, create a new one
except FileNotFoundError:
    wb = Workbook()

model2.write_schedule(wb)
date = 1
wb.save('production_schedule_{}.xlsx'.format(date))
# get the last molds to use as input for next weeks scheduling
next_mount = model2.get_last_molds()
