# -*- coding: utf-8 -*-
"""
Created on Wed Apr  1 10:14:05 2020

@author: Ook
"""

from openpyxl.styles import Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


wb = Workbook()
ws = wb.active

c = ws['B3']
c.value = 'abcdef'


def set_border(ws, row_start=None, col_start=None, row_end=None, col_end=None, weight='thin', outer=None):
    side = Side(border_style=weight)
    # if no start row/col, return
    if not row_start or not col_start:
        return
    # allow for single cell, row, column
    # if row/col end not given or smaller than start row/col
    if not row_end or row_end < row_start:
        row_end = row_start
    if not col_end or col_end < col_start:
        col_end = col_start
        
    for i in range(row_start,row_end+1):
        for j in range(col_start,col_end+1):
            if i == row_start or i == row_end or j == col_start or j == col_end:
                c = ws.cell(i,j)
                border = Border(c.border.left,c.border.right,c.border.top,c.border.bottom)
                if i == row_start:
                    border.top = side
                if i == row_end:
                    border.bottom = side
                if j == col_start:
                    border.left = side
                if j == col_end:
                    border.right = side
                c.border = border
        
    
    
def fit_column(ws,row_start,col,row_end, min_width=8.43):
    width = max(len(as_text(ws.cell(i,col).value)) for i in range(row_start, row_end+1))+1
    ws.column_dimensions[get_column_letter(ws.cell(row_start,col).column)].width = max(width,min_width)
    
def as_text(value):
    return str(value) if value is not None else ''
    
#wb.save('schedule.xlsx')