"""A couple of classes that model and optimize production
for Alpha Systems.
"""

import gurobipy as gp
from gurobipy import GRB
import pandas as pd
import deep_ordered_dict as dod
from workbook_utils import set_border, fit_column
from openpyxl.styles import Alignment
from datetime import datetime

class BaseClass:    # probably needs a better name
    def __init__(self, hours, arms, mounts, name):
        """Initialize production capacity by setting the number of arms,
        mounts, and production hours.

        Parameters
        ----------
        hours: int--length of production run
        arms: int--number of total arms available
        mounts: int--number of mounts per arm
        name: str--name of the sover model
        """
        self.hours = hours
        self.arms = arms
        self.mounts = mounts
        self.model = gp.Model(name=name)
        self.sched = None
        
    def __getattr__(self, attr):
        return getattr(self.model,attr)
    
    def optimize(self, term_conds):
        """Run the optimize method of the model with termination conditions
        
        Iterate through the list of (tolerance,time_limit) variables and 
        assign them to the solver's tolerance and time limit, then run the solver.
        
        Example
        -------
        Let term_conds = [(0,100),(0.5,200)], then the solver will terminate
        if abs(lb-ub) = 0%*abs(ub) and solve time < 100 secs,
        or 0 <= abs(lb-ub) <= 0.5%*abs(ub) and 
        100 secs < solve time < 200 secs, or solve time >= 200 secs.
        
        Parameters
        ----------
        term_conds: list-like 
            a list-like object of (tolerance, time_limit) number tuples
        """
        for tolerance, time_limit in term_conds:
            self.model.Params.MIPGap = tolerance
            self.model.Params.timelimit = time_limit
            self.model.optimize()
            # if terminate by timelimit, then optimize again with next pair
            # of arguments, else quit loop.
            if self.model.status != 9:
                break
            
    def read_data(self,product_data, mold_data):
        """Reads product and mold data needed for building the linear program model

        Seperate product data frame into series by column label and assign them as attributes to the object.
        Assign the mold data as attribute of the object.
        
        Parameters
        ----------
        part_data: pd.DataFrame
            a data frame indexed (row) by part number(string) with columns
            demand (int), mold (string). all numbers are non negative.
        mold_data: pd.Series
            a series indexed (row) by mold number (string) with value quantity of mold (int).
        
        Notes
        -----
        Assumes that the the mold numbers in the index of mold_data are the same as
        in the 'mold' column of product_data.
        """
        # need a copy of product_data to use masking to iterate over 
        # the part number for each mold
        self.data = product_data
        self.parts = self.data.index
        self.demands = self.data['demand']
        self.molds = mold_data.index
        self.qty_molds = mold_data['qty_mold']
        
    def reset(self):
        self.model.reset()

    def get_production(self):
        # returns a pd.Series of production indexed by part number
        return self.prods
    
    def get_schedule(self):
        return self.sched
    
    def is_optimized(self):
        return self.model.status in (2,9)

class MaxProfit(BaseClass):
    """

    """

    def __init__(self, hours, arms, mounts):
        BaseClass.__init__(self,hours,arms,mounts,'Maximize_Profit')
        
    def read_data(self, product_data, mold_data):
        """see BaseClass.read_data
        
        The difference is this method requires columns inv (int),
        profit (float), desired (int) for product_data
        """
        BaseClass.read_data(self,product_data,mold_data)
        self.inv = self.data['inv']
        self.profits = self.data['profit']
        self.desired = self.data['desired']
        
    def build_model(self):
        """Build a profit maximization linear program from the attributes 
        created by self.read_data.

        Add variables, constraints and objective function to self.model, using the
        values from 'data' to create a linear program that determines production for
        each product to maximize total profit while meeting each product's demand.

        Notes
        -----
        This method only builds the model without optmizing it.
        i is used to subscript part number and j for mold number.
        """
        # prep data
        # max(inv,desired,1)
        desired = self.data[['inv','desired']].max(axis=1).clip(lower=1)
        # or desired = pd.DataFrame(self.inv,self.desired).max().clip(lower=1)
        # or desired = pd.concat([self.inv,self.desired],axis=1).max(axis=1).clip(lower=1)
        
        # some general variables
        max_molds = self.qty_molds.max()
        max_cap = self.hours*self.arms*self.mounts

        # create decision variables
        self.pro_vars = self.model.addVars(self.parts,vtype=GRB.INTEGER,name='pro_vars')
        self.a = self.model.addVars(self.parts,vtype=GRB.BINARY,name='a')
        # b and c by mold number since they concern mold changing
        self.b = self.model.addVars(self.molds,vtype=GRB.BINARY,name='b')
        self.c = self.model.addVars(self.molds,vtype=GRB.BINARY,name='c')

        # Big-M number for meeting demand
        BIG_M1 = self.demands.max()
        # Big-M number for objective funciton penalty for not meeting demand
        BIG_M2 = 100000 #2*self.hours*self.profits.max()
        
        # constraints
        # the sum of productions for all the parts that use a certain mold <= hours * available quantity of that mold
        self.c1 = self.model.addConstrs((gp.LinExpr((1,self.pro_vars[i]) for i in self.data[self.data['mold']==j].index) <= self.hours*self.qty_molds[j] for j in self.molds),name='const1')
        # sum of production + (number of mounts per arm * molds with production between 1-23) <= max cap
        # note that this is always true only if hours >= 24 + mounts
        self.c2 = self.model.addConstr((gp.LinExpr((1,self.pro_vars[i]) for i in self.parts) + gp.quicksum(self.mounts - self.mounts*self.c[j] for j in self.molds) <= max_cap),name='const2')
        # if production + inventory < demand, then a = 1
        self.c3 = self.model.addConstrs((self.pro_vars[i] + self.inv[i] + BIG_M1*self.a[i] >= self.demands[i] for i in self.parts),name='const3')
        # production + inventory <= desired
        self.c4 = self.model.addConstrs((self.pro_vars[i] + self.inv[i] - self.demands[i] <= desired[i] for i in self.parts),name='const4')
        # if production of a mold (all parts that use the mold) < 24, b = 0
        self.c5 = self.model.addConstrs((24*self.b[j] <= gp.LinExpr((1,self.pro_vars[i]) for i in self.data[self.data['mold']==j].index) for j in self.molds),name='const5')
        # if production of a mold is between 1-23, c = 0
        self.c6 = self.model.addConstrs((gp.LinExpr((1,self.pro_vars[i]) for i in self.data[self.data['mold']==j].index) <= 24 + max_molds*self.hours*self.b[j] - 24*self.c[j] for j in self.molds),name='const6')

        # objective function
        # sum of profit*(production + inventory - demand) with penalties for each product not meeting demand
        self.obj = self.model.setObjective(gp.quicksum((self.profits[i]*(self.pro_vars[i]+self.inv[i]-self.demands[i])-BIG_M2*self.a[i]) for i in self.parts),GRB.MAXIMIZE)

    def update_production(self):
        """if the model is optimized ,create pd.Series of the production deteremined by the model (optimal solution)
        indexed by part numbers and assigns to self.prods.
        """
        if self.is_optimized():
            self.prods = pd.Series({i:self.pro_vars[i].x for i in self.parts},name='produced')
            return True
        else:
            # do something
            return False
        
            

class ProductionSchedule(BaseClass):
    """

    """
    def __init__(self, hours, arms, mounts):
        BaseClass.__init__(self,hours,arms,mounts,'Minimize_Mold_Change')

    def read_data(self, product_data, mold_data, prev_mount=None):
        """see BaseClass.read_data
        
        The difference is this method requires column produced (int),
        for product_data

        Parameters
        ----------
        prev_mount: pd.Series indexed by arm:mount and with value
            mold number mounted. (e.g., prev_mount['1:2'] == 'the_mold_number'.)
            These are the last molds (on each mount) mounted from the previous production run.
        """
        BaseClass.read_data(self,product_data,mold_data)
        self.prods = self.data['produced']
        self.prev_mount = prev_mount
        
    def build_phase1(self):
        """Build a production maximization linear program from the attributes 
        created by self.read_data.

        Add variables, constraints and objective function to self.model, using the
        values from 'data' to create a linear program that determines production at each 
        arm/mount for each product to maximize total production subject to the production
        of each part <= desired production while meeting each product's demand.
        
        Notes
        -----
        This method only builds the model without optmizing it.
        i is used to subscript arm number, j for mount number, k for part number,
        and h for mold number.
        """
        
        # prep data
        # min(demand,prod) by part number
        demands = self.data[['demand','produced']].min(axis=1)
        # or demands = pd.DataFrame([self.demands,self.prods]).min()
        # or demands = pd.concat([self.demands,self.prods],axis=1).min(axis=1)

        # some general variables
        arms = [x for x in range(self.arms)]    # so dont have to call range(self.arms) all the time
        mounts = [x for x in range(self.mounts)]
        BIG_M1 = 10000
        BIG_M2 = 1000000
        max_cap = self.hours*self.arms*self.mounts

        # create decision variables
        self.pro_vars = self.model.addVars(arms,mounts,self.parts,vtype=GRB.INTEGER,name='pro_vars')
        # a[i,j,k] keeps tracks of production of part k on mount j arm i
        self.a = self.model.addVars(arms,mounts,self.parts,vtype=GRB.BINARY,name='a')
        self.b = self.model.addVars(arms,mounts,vtype=GRB.BINARY,name='b')
        self.c = self.model.addVars(self.parts,vtype=GRB.BINARY,name='c')
        # a1[i,j,h] keeps track of production on mold h mount j arm i
        # the difference is between a[i,j,k] and a1[i,j,h] that there can be more than one part that use the same mold
        self.a1 = self.model.addVars(arms,mounts,self.molds,vtype=GRB.BINARY,name='a1')

        # create constraints
        # c[k] = 1 if pro_vars[k] < demand [k]
        self.c1 = self.model.addConstrs((self.pro_vars.sum('*','*',k) + BIG_M1*self.c[k] >= demands[k] for k in self.parts), name='const1')

        # a[i,j,k] = 1 if pro_vars[i,j,k] >= 1
        self.c2 = self.model.addConstrs((self.pro_vars[i,j,k] - self.hours*self.a[i,j,k] <= 0 for i in arms for j in mounts for k in self.parts), name='const2')
        # a[i,j,k] = 0 if pro_vars[i,j,k] = 0
        #self.c3 = self.model.addConstrs((self.pro_vars[i,j,k] - self.a[i,j,k] >= 0 for i in arms for j in mounts for k in self.parts), name='const3')

        # a1[i,j,h] = 1 if the sum of (a[i,j,k] such that part k uses mold h) >= 1
        self.c3 = self.model.addConstrs((gp.LinExpr((1,self.a[i,j,k]) for k in self.data[self.data['mold']==h].index) <= self.hours*self.a1[i,j,h] for i in arms
                                         for j in mounts for h in self.molds),name='const3')
        # a1[i,j,h] = 0 if the sum of (a[i,j,k] such that part k uses mold h) = 0
        """self.c4 = self.model.addConstrs((gp.LinExpr((1,self.a[i,j,k]) for k in self.data[self.data['mold']==h].index) >= self.a1[i,j,h] for i in arms
                                         for j in mounts for h in self.molds),name='const4')"""
        
        # b[i,j] = 1 if a1[i,j,'*'] >= 1
        self.c5 = self.model.addConstrs((self.a1.sum(i,j,'*') - self.hours*self.b[i,j] <= 0 for i in arms for j in mounts), name='const5')
        # b[i,j] = 0 if a1[i,j,'*'] = 0
        self.c6 = self.model.addConstrs((self.a1.sum(i,j,'*') - self.b[i,j] >= 0 for i in arms for j in mounts), name='const6')

        # production + mold changes <= hours 
        self.c7 = self.model.addConstrs((self.pro_vars.sum(i,j,'*') + self.a1.sum(i,'*','*') - self.b.sum(i,'*') <= self.hours
                                   for i in arms for j in mounts), name='const7')
        # pro_vars[k] <= prods[k]
        self.c8 = self.model.addConstrs((self.pro_vars.sum('*','*',k) <= self.prods[k] for k in self.parts), name='const8')
        # number of molds mounted <= available molds
        self.c9 = self.model.addConstrs((self.a1.sum('*','*',h) <= self.qty_molds[h] for h in self.molds), name='const9')
        

        # fix a1[i,j,h] == 1 if prev_mount['i:j'] = h
        # might need to check that the 'arm:mount' keys are consistent with the built moldel
        self.c10 = self.model.addConstrs((self.a1[int(idx.split(':')[0]),int(idx.split(':')[1]),self.prev_mount[idx]] == 1 for idx in self.prev_mount.index), name='const10')
        
        # set objective function
        self.obj = self.model.setObjective(self.pro_vars.sum()-BIG_M2*self.c.sum(),GRB.MAXIMIZE)

    def build_phase2(self):
        """Changes self.model to a mold change minimization linear program from
        the attributes created by self.read_data.
        
        Minimizes mold changes subject to the production of each part == the 
        production determined by optimizing the phase1 model (i.e., minimize 
        the mold changes while retaining maximum production).
        Also group the production of each part together as much as possible.
        (e.g., suppose two products use the same two molds, the model will try
        have each mold making a product, instead of having both molds do half
        of each product.)
        
        Notes
        -----
        This method only builds the model without optimizing it.
        The method uses the return value of self.get_production() to ensure a feasible 
        solution (self.c11).
        """
        if self.is_optimized():
            # update production (self.prods) to the result from phase 1
            self.update_production()
            # remove obsolete constraint
            self.model.remove(self.c8)
            # production for each part = to those from phase1's optimal solution
            self.c11 = self.model.addConstrs((self.pro_vars.sum('*','*',k) == self.prods[k] for k in self.parts), name='const11')
            # set objective
            # this works because minimizing sum(all a1) doesn't increase sum(all a)
            # because for each a[i,j,k] = 1, a[i,k,h] also = 1, where part k uses mold h
            self.obj2 = self.model.setObjective(self.a.sum() + self.a1.sum() - self.b.sum(),GRB.MINIMIZE)
            # incase the obove assertion isn't true, we can use
            # self.obj2 = self.model.setObjective(self.a.sum() + BIG-M*(self.a1.sum() - self.b.sum()),GRB.MINIMIZE)
        
    def update_production(self):
        """Store production by mold number in a pd.Series
        might have to change to by part number production
        """
        if self.is_optimized():
            self.prods = pd.Series({k:self.pro_vars.sum('*','*',k).getValue() for k in self.parts}, name='produced')
            return True
        else:
            return False
        
    def update_schedule(self):
        """Store production distribution for each part on 
        each arm/mount in nested DeepOrderedDict and assign to self.sched

        Take the current optimal solution and store the values in nested DeepOrderedDicts.
        Call self.move_to_top to rearrange the schedule.
        
        Notes
        -----
        if called when phase2 is not optimized yet, then the schedule may not 
        have minimum mold changes
        """
        if self.is_optimized():
            outer = dod.DeepOrderedDict()
            for i in range(self.arms):
                for j in range(self.mounts):
                    # if there is production on this mount
                    if self.pro_vars.sum(i,j,'*').getValue():
                        key = '{}:{}'.format(i,j)
                        outer[key] = dod.DeepOrderedDict()
                        for h in self.molds:
                            # if mold is mounted
                            if self.a1[i,j,h].x == 1:
                                outer[key][h] = dod.DeepOrderedDict([(k,int(self.pro_vars[i,j,k].x)) for k in self.data[self.data['mold']==h].index if self.pro_vars[i,j,k].x >0])
            self.sched = outer
            self.move_to_top()
            return True
        else: 
            return False
        
    def move_to_top(self):
        """move molds that are mounted last from previous session to
        the top of their respective mounts.
        """
        for mount in self.prev_mount.index:
            self.sched[mount].move_to_end(self.prev_mount[mount],last=False)
    
    def get_last_molds(self):
        """Return a pd.Series indexed by the mount number with the values being
        the mold number of the last mold mounted on each mount.
        """
        if not self.sched:
            return None
        cur_mount = dict()
        for mt in self.sched:
            cur_mount[mt] = self.sched[mt].get_end(what='key',depth=0)            
        return pd.Series(cur_mount)
    
    def get_mold_changes(self):
        return int((self.a1.sum()-self.b.sum()).getValue())

    def write_schedule(self,wb,week=1,start_time=datetime.now().replace(second=0,microsecond=0)):
        """write the production schedule to wb openpyxl Workbook
        """
        # check if schedule available
        if not self.sched:
            # if self.sched is None (no real schedule), return False
            return False
        # function start
        old_time = start_time
        current_time = old_time
        ws = wb.create_sheet('week{}'.format(week))
        align_center = Alignment('center','center')
        
        # start entries on row 3, 1st and 2nd for titles, etc.
        # 1st column for date, 2nd for time

        # write first date in first entry
        ws.cell(row=3,column=1,value=start_time.date())
        # write time in all rows in 2nd column,
        for i in range(3,self.hours+3):
            ws.cell(row=i,column=2,value=current_time.time())
            current_time = old_time + pd.Timedelta(hours=1)
            # if date changes (next day), write date in 1st column
            if current_time.date() != old_time.date():
                ws.cell(row=i,column=1,value=current_time.date())
            old_time = current_time
        # set border and fit date/time column
        set_border(ws,3,1,self.hours+2)
        set_border(ws,3,2,self.hours+2)
        fit_column(ws,3,1,self.hours+2)
        fit_column(ws,3,2,self.hours+2)
        # set border over top of schedule
        set_border(ws,2,3,2,self.arms*self.mounts+2)

        # enter productions
        # first column of production entries
        col = 3
        for mt in self.sched:
            # keys of self.sched are mount numbers
            mount = self.sched[mt]
            # enter column (mount) label
            cell = ws.cell(2,col,mt)
            cell.alignment = align_center
            # start at row 3
            row = 3
            for h in mount:
                # keys of mount are mold numbers
                mold = mount[h]     
                for k in mold:
                    # keys of mold are part numbers
                    cell = ws.cell(row,col,k)
                    cell.alignment = align_center
                    # if production of part k > 1,
                    # enter production in the cell below
                    if mold[k] > 1:
                        ws.cell(row+1,col,mold[k])
                    # set border for production of current part
                    set_border(ws,row,col,row+mold[k]-1)
                    # set row to current row + production hours of part
                    row += mold[k]  
                # skip one row for mold change
                row += 1
            # fit column width >= widest value in the column
            fit_column(ws,3,col,self.hours+2)
            # move col to the next column on the right
            col += 1
        # add freez panse 
        ws.freeze_panes = 'C3'

        return True

    


        
